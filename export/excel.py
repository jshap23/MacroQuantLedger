from __future__ import annotations
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from models.schema import AppState

# ── Palette ────────────────────────────────────────────────────────────────────
NAVY      = "1B2A4A"
GOLD      = "C9A84C"
WHITE     = "FFFFFF"
CREAM     = "FDF8F0"
DEEP_NAVY = "0f1a2e"

# ── Rank score colors (AssetView direction = "1"–"5" or "—") ──────────────────
RANK_BG = {
    "5": "14532d", "4": "166534", "3": "374151",
    "2": "7c2d12", "1": "7f1d1d", "—": "2a2a38",
}
RANK_FG = {
    "5": "4ade80", "4": "86efac", "3": "9ca3af",
    "2": "fb923c", "1": "f87171", "—": "666688",
}

# ── Direction colors (MacroView direction = Bullish/Neutral/Bearish/No View) ───
DIR_BG = {
    "Bullish": "1a6b3a", "Neutral": "2a2a38",
    "Bearish": "6b1a1a", "No View": "1e1e24",
}
DIR_FG = {
    "Bullish": "4ade80", "Neutral": "a0a0b8",
    "Bearish": "f87171", "No View": "555566",
}


# ── Style helpers ──────────────────────────────────────────────────────────────

def _hdr_font():    return Font(name="Calibri", bold=True, color=WHITE, size=11)
def _body_font():   return Font(name="Calibri", size=10)
def _bold_font():   return Font(name="Calibri", bold=True, size=10)
def _navy_fill():   return PatternFill("solid", fgColor=NAVY)
def _cream_fill():  return PatternFill("solid", fgColor=CREAM)
def _white_fill():  return PatternFill("solid", fgColor=WHITE)
def _gold_fill():   return PatternFill("solid", fgColor=GOLD)
def _deep_fill():   return PatternFill("solid", fgColor=DEEP_NAVY)


def _border(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _set_col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_headers(ws, row: int, headers: list[str], height: int = 20):
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = _hdr_font()
        c.fill = _navy_fill()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border()
    ws.row_dimensions[row].height = height


def _section_title(ws, row: int, title: str, ncols: int):
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(name="Calibri", bold=True, color=WHITE, size=11)
    c.fill = _gold_fill()
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    if ncols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.row_dimensions[row].height = 22


def _fmt_dt(dt) -> str:
    if dt is None:
        return "—"
    if hasattr(dt, "strftime"):
        return dt.strftime("%Y-%m-%d")
    return str(dt)


def _alt_fill(i: int):
    return _cream_fill() if i % 2 else _white_fill()


# ── Sheet 1: Macro Views ───────────────────────────────────────────────────────

def _sheet_macro_views(wb, state: AppState):
    ws = wb.create_sheet("Macro Views")
    NCOLS = 9
    headers = ["View", "Variable", "Directional Lean",
               "Signal 1", "Signal 2", "Signal 3",
               "The Counter", "Conviction", "Last Touched"]
    _write_headers(ws, 1, headers)
    ws.freeze_panes = "B2"

    for i, v in enumerate(state.macro_views):
        r = i + 2
        fill = _alt_fill(i)
        values = [
            v.direction,
            v.name,
            v.lean,
            v.signals[0] if len(v.signals) > 0 else "",
            v.signals[1] if len(v.signals) > 1 else "",
            v.signals[2] if len(v.signals) > 2 else "",
            v.counter,
            v.conviction,
            _fmt_dt(v.last_touched),
        ]
        for col, val in enumerate(values, start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = _body_font()
            c.fill = fill
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = _border()

        # Color-code the "View" direction badge
        dc = ws.cell(row=r, column=1)
        bg = DIR_BG.get(v.direction, DIR_BG["No View"])
        fg = DIR_FG.get(v.direction, DIR_FG["No View"])
        dc.fill = PatternFill("solid", fgColor=bg)
        dc.font = Font(name="Calibri", bold=True, color=fg, size=10)
        dc.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[r].height = 55

    # Macro notes footer
    if state.macro_notes:
        nr = len(state.macro_views) + 3
        lbl = ws.cell(row=nr, column=1, value="CROSS-CUTTING NOTES")
        lbl.font = Font(name="Calibri", bold=True, color=GOLD, size=10)
        lbl.fill = _deep_fill()
        lbl.alignment = Alignment(vertical="center", indent=1)
        ws.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=NCOLS)
        ws.row_dimensions[nr].height = 18
        nr += 1
        nc = ws.cell(row=nr, column=1, value=state.macro_notes)
        nc.font = _body_font()
        nc.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=NCOLS)
        ws.row_dimensions[nr].height = 80

    _set_col_widths(ws, [13, 18, 40, 26, 26, 26, 32, 12, 14])


# ── Sheet 2: Asset Class Views ─────────────────────────────────────────────────

def _sheet_asset_views(wb, state: AppState):
    """
    Layout per group (L1 / Equities / Fixed Income):
      [SECTION HEADER]
      [Col headers: Asset | Rank | Conviction | One-Liner Thesis | Last Touched]
      For each asset:
        Asset row  — rank cell is color-coded by score (1=red → 5=green)
        Deep Dive  — indented sub-row in deep-navy if commentary exists
      [blank spacer row]
    """
    ws = wb.create_sheet("Asset Class Views")
    NCOLS = 4
    headers = ["Asset", "Rank", "One-Liner Thesis", "Last Touched"]

    groups = [
        ("l1",           "LEVEL 1 — CROSS-ASSET"),
        ("equities",     "EQUITIES"),
        ("fixed_income", "FIXED INCOME"),
    ]

    row = 1
    ws.freeze_panes = "A2"

    for group_key, group_label in groups:
        views = [v for v in state.asset_views if v.group == group_key]
        if not views:
            continue

        _section_title(ws, row, group_label, NCOLS)
        row += 1
        _write_headers(ws, row, headers, height=18)
        row += 1

        for i, av in enumerate(views):
            fill = _alt_fill(i)

            # ── Primary data row ──────────────────────────────────────────────
            values = [av.name, av.direction, av.note, _fmt_dt(av.last_touched)]
            for col, val in enumerate(values, start=1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = _body_font()
                c.fill = fill
                c.alignment = Alignment(vertical="center", wrap_text=True)
                c.border = _border()
            ws.row_dimensions[row].height = 26

            # Color-code the Rank cell
            rank_cell = ws.cell(row=row, column=2)
            score = av.direction

            rank_cell.fill = PatternFill("solid", fgColor=RANK_BG.get(score, RANK_BG["—"]))
            rank_cell.font = Font(name="Calibri", bold=True,
                                  color=RANK_FG.get(score, RANK_FG["—"]), size=12)
            rank_cell.alignment = Alignment(horizontal="center", vertical="center")

            # Make asset name bold
            ws.cell(row=row, column=1).font = _bold_font()

            row += 1

            # ── Deep dive sub-row (only if commentary exists) ─────────────────
            if av.commentary and av.commentary.strip():
                # Label cell
                lbl = ws.cell(row=row, column=1, value="DEEP DIVE")
                lbl.font = Font(name="Calibri", bold=True, color=GOLD, size=8)
                lbl.fill = _deep_fill()
                lbl.alignment = Alignment(horizontal="left", vertical="top", indent=2)
                lbl.border = _border("333355")

                # Commentary cell (merged across cols 2–5)
                txt = ws.cell(row=row, column=2, value=av.commentary)
                txt.font = Font(name="Calibri", size=9, color="bbbbcc", italic=True)
                txt.fill = _deep_fill()
                txt.alignment = Alignment(vertical="top", wrap_text=True)
                txt.border = _border("333355")
                ws.merge_cells(start_row=row, start_column=2,
                               end_row=row, end_column=NCOLS)


                # Estimate row height from commentary length
                est_lines = max(2, len(av.commentary) // 80 + 1)
                ws.row_dimensions[row].height = min(est_lines * 14, 120)
                row += 1

        # Blank spacer row between groups
        ws.row_dimensions[row].height = 10
        row += 1

    _set_col_widths(ws, [24, 7, 60, 14])



def _write_row(ws, row: int, values: list, alt: bool = False):
    fill = _cream_fill() if alt else _white_fill()
    for col, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=col, value=v)
        c.font = _body_font()
        c.fill = fill
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = _border()
    ws.row_dimensions[row].height = 42


# ── Sheet 5: Weekly Reconciliation ────────────────────────────────────────────

def _sheet_reconciliation(wb, state: AppState):
    ws = wb.create_sheet("Weekly Reconciliation")
    headers = ["Week Of", "Macro Scan", "Quant Check",
               "% Macro", "% Quant", "% Other", "Synthesis"]
    _write_headers(ws, 1, headers)
    ws.freeze_panes = "A2"

    for i, r in enumerate(state.reconciliations):
        fill = _alt_fill(i)
        values = [
            r.date.strftime("%Y-%m-%d") if r.date else "—",
            r.macro_scan,
            r.quant_check,
            r.time_macro,
            r.time_quant,
            r.time_other,
            r.synthesis,
        ]
        for col, val in enumerate(values, start=1):
            c = ws.cell(row=i + 2, column=col, value=val)
            c.font = _body_font()
            c.fill = fill
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = _border()

        # Center-align the percentage columns
        for col in range(4, 7):
            ws.cell(row=i + 2, column=col).alignment = Alignment(
                horizontal="center", vertical="center"
            )
        ws.row_dimensions[i + 2].height = 65

    _set_col_widths(ws, [14, 46, 46, 9, 9, 9, 50])


# ── Public entry point ─────────────────────────────────────────────────────────

def generate_excel(state: AppState) -> Path:
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    _sheet_macro_views(wb, state)
    _sheet_asset_views(wb, state)
    _sheet_reconciliation(wb, state)

    date_str = datetime.now().strftime("%Y-%m-%d")
    out_dir = Path(__file__).parent.parent / "data"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"macroquant_ledger_{date_str}.xlsx"
    wb.save(out_path)
    return out_path
